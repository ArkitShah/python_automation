# In this project, we have a price column in which the prices are wrongly entered. We need to add the corrected price
# which is 10% less from the original price. For that we will use the openpyxl library which allows us to access and
# modify the contents in an excel spreadsheet. At the end it also generated a Bar Chart from the new data.

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# Creating a function 'process_workbook' which takes input file name and output file name as parameters.

def process_workbook(input_filename, output_filename):
    wb = xl.load_workbook(input_filename)
    sheet = wb['Sheet1']

    # Traversing through the rows.
    # Starting from 2nd row as 1st row contains the labels.

    for row in range(2, sheet.max_row + 1):
        # Accessing the 3rd column i.e. price.

        cell = sheet.cell(row, 3)

        # Making a new column which contains the corrected price.

        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # Creating an instance 'chart' of class 'BarChart'.

    chart = BarChart()
    chart.add_data(values)

    # Creating the Bar Chart at cell 'g2' by taking values from the new column i.e. corrected price column.

    sheet.add_chart(chart, 'g2')
    wb.save(output_filename)


process_workbook('transactions.xlsx', 'transactions2.xlsx')
